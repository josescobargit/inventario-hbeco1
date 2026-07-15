import csv
import io
from dataclasses import dataclass


@dataclass(frozen=True, slots=True)
class ParsedCount:
    row: int
    sku: str
    physical_confirmed: int


@dataclass(frozen=True, slots=True)
class ImportError:
    row: int | None
    column: str | None
    sku: str | None
    message: str


def parse_stock_csv(content: bytes) -> tuple[list[ParsedCount], list[ImportError]]:
    try:
        text = content.decode("utf-8-sig")
    except UnicodeDecodeError:
        return [], [ImportError(None, None, None, "Guarda el archivo como CSV UTF-8.")]
    reader = csv.DictReader(io.StringIO(text))
    required = {"SKU", "Stock_Fisico"}
    if not reader.fieldnames or not required.issubset(set(reader.fieldnames)):
        return [], [
            ImportError(1, None, None, "Se esperan las columnas SKU y Stock_Fisico.")
        ]

    parsed: list[ParsedCount] = []
    errors: list[ImportError] = []
    seen: set[str] = set()
    for row_number, row in enumerate(reader, start=2):
        sku = (row.get("SKU") or "").strip().upper()
        raw_quantity = (row.get("Stock_Fisico") or "").strip()
        if not sku:
            errors.append(
                ImportError(row_number, "SKU", None, "El SKU es obligatorio.")
            )
            continue
        if sku in seen:
            errors.append(ImportError(row_number, "SKU", sku, "El SKU está duplicado."))
            continue
        seen.add(sku)
        try:
            quantity = int(raw_quantity)
            if str(quantity) != raw_quantity or quantity < 0:
                raise ValueError
        except ValueError:
            errors.append(
                ImportError(
                    row_number,
                    "Stock_Fisico",
                    sku,
                    "Debe ser un entero mayor o igual a cero.",
                )
            )
            continue
        parsed.append(ParsedCount(row_number, sku, quantity))
    return parsed, errors


def parse_stock_xlsx(content: bytes) -> tuple[list[ParsedCount], list[ImportError]]:
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True))
    except Exception:
        return [], [
            ImportError(
                None,
                None,
                None,
                "No pudimos leer el archivo XLSX. Descarga una plantilla nueva.",
            )
        ]
    if not values:
        return [], [ImportError(1, None, None, "El archivo está vacío.")]
    headers = [str(value or "").strip() for value in values[0]]
    if "SKU" not in headers or "Stock_Fisico" not in headers:
        return [], [
            ImportError(1, None, None, "Se esperan las columnas SKU y Stock_Fisico.")
        ]
    sku_index, stock_index = headers.index("SKU"), headers.index("Stock_Fisico")
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["SKU", "Stock_Fisico"])
    for row in values[1:]:
        writer.writerow(
            [
                row[sku_index] if sku_index < len(row) else "",
                row[stock_index] if stock_index < len(row) else "",
            ]
        )
    return parse_stock_csv(output.getvalue().encode("utf-8"))
